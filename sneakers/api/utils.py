"""

Prisma Inc.

utils.py

Status: UNDER DEVELOPMENT for Multiprocessing Implementation

Notes:
        download_images(df):
        Works but uses the multiprocessing system so, under dev.

Made by Alexis W.

"""
import pandas as pd
import os
import requests
import progressbar
import gc
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyImage
from sneakers.api import processing
from datetime import datetime

def load_shoes_dataset_quantity(q):
    shoes = pd.read_csv('sneakers/datasets/sneakers-local-db.csv', index_col=0, low_memory=False)
    out = shoes.iloc[:q]
    return out

# Status: Check
def load_shoes_dataset():
    shoes = pd.read_csv('sneakers/datasets/sneakerstempdb2live.csv', low_memory=False)
    return shoes


# Status: Check
def flush_sheets():

    arr = os.listdir('sheets')

    for file in arr:
        os.remove('sheets/{fname}'.format(fname=file))

    f = open("sheets/phasm.txt", "w+")
    f.write('Prisma Inc. 2021')
    f.close()

    return True


def flush_airsheets():

    arr = os.listdir('airout')

    for file in arr:
        os.remove('airout/{fname}'.format(fname=file))

    f = open("airout/phasm.txt", "w+")
    f.write('Prisma Inc. 2021')
    f.close()

    return True


# Status: Check
def flush_outputs():

    arr = os.listdir('outputs')

    for file in arr:
        os.remove('outputs/{fname}'.format(fname=file))

    return True


# Status: Check
def flush_img():

    arr = os.listdir('img')

    for file in arr:
        os.remove('img/{fname}'.format(fname=file))

    f = open("img/phasm.txt", "w+")
    f.write('Prisma Inc. 2021')
    f.close()

    return True


# Status: Check
def build_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df.to_excel(path)

    wb = load_workbook(filename=path)

    ws = wb.active

    progress_lenc = len(df.index)

    progsc = progressbar

    for i in progsc.progressbar(range(0, progress_lenc)):

        TestUrl = df['image'][i]

        cellName = 'S{fnum}'.format(fnum=i + 2)

        try:

            im = Image.open(requests.get(TestUrl, stream=True).raw)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(i + 2)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

        except:

            continue

    wb.save(path)
    print('Images downloaded succesfully...')
    print('XLSX builded...')
    print('Check /sheets folder for {ffile}'.format(ffile = path))

    return wb


# Status: Check
def build_large_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df['pic'] = ''

    writer = pd.ExcelWriter(path, engine='xlsxwriter', options={'strings_to_urls': False})
    df.to_excel(writer)
    writer.close()

    wb = load_workbook(filename=path)

    progress_lenc = len(df.index)

    progsc = progressbar

    for i in progsc.progressbar(range(0, progress_lenc)):

        ws = wb.active

        TestUrl = df['image'][i]

        j = i + 2

        cellName = 'S{fnum}'.format(fnum=j)

        try:

            im = Image.open(requests.get(TestUrl, stream=True).raw)

            # time.sleep(3)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(j)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

            wb.save(path)

            del img
            del xImg
            del cellName
            del TestUrl
            del loc
            del im_100
            del im

            gc.collect()

        except requests.exceptions.MissingSchema:

            #print('MissingSchema Error on iteration number {fit}'.format(fit=j))

            continue

        except requests.exceptions.ConnectionError:

            #print('Connection Error on iteration number {fit}'.format(fit=j))

            continue

    print('Images downloaded succesfully...')
    print('XLSX large builded...')
    print('Check /sheets folder for {ffile}'.format(ffile=path))

    return True


# Status: Checked
def download_img(df):

    ver = 'image-download-dummy-file-delete-this'

    title = "{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    #df.to_excel(path)

    processing.img_download_processing(df, path)

    print('Images downloaded succesfully')


# Status: For Revision
def build_big_xlsx(df, ver, local=False):

    # MULTI-PROCESSING MODULE

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df['pic'] = ''

    writer = pd.ExcelWriter(path, engine='xlsxwriter', options={'strings_to_urls': False})
    df.to_excel(writer)
    writer.close()

    #wb = load_workbook(filename=path)

    #ws = wb.active

    if local is True:
        processing.processing_xlsx_local(df, path)
    else:
        processing.processing_xlsx(df, path)


    print('Images downloaded succesfully...')
    print('XLSX large builded...')
    print('Check /sheets folder for {ffile}'.format(ffile=path))

    return True


# DEVELOPMENT SECTION
def build_xlxs_injector(df, ver, size, local=True):

    # MULTI-PROCESSING MODULE

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df['pic'] = ''

    # Creates the Excel Worksheet for working
    writer = pd.ExcelWriter(path, engine='xlsxwriter', options={'strings_to_urls': False})
    df.to_excel(writer)
    writer.close()
    # Excel Worksheet created in the specified 'path'


    # Sends 'worksheet' and 'dataframe' to processing
    if local is True:
        # This is the process we are going to work for the injection
        processing.processing_xlsx_local_inj(df, path, size)
        print('DEV: BUILD BY INJECTION PROCESS ENDED')
    else:
        processing.processing_xlsx_inj(df, path, size)

    print('XLSX large build...')
    print('Check /sheets folder for {ffile}'.format(ffile=path))

    return True


def sheet_ver(title, sub, sp):

    # Get current datetime
    now = datetime.now()

    # Change the date format
    nowstr = now.strftime("%d-%m-%Y %H-%M-%S")


    # Set up the title for the xlsx
    ver = '{ftitle}({fsub})-({flen}rows)-{ftime}'.format(ftitle=title,ftime=nowstr, flen=sp, fsub=sub)

    return ver


def composer_title(title, sub):

    return '{ftitle}({fsub})'.format(ftitle=title, fsub=sub)




