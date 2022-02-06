"""

Prisma Inc. 2021

processing.py

Status: Checked

Note: Multi-processing algorithms for faster image manipulation.

Made by Alexis W.

"""
import requests
import progressbar
import time
import gc
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyImage
from multiprocessing import Pool
from sneakers.api import injector
from sneakers.vision import dataset


# Basic XLSX process
def img_process(df, path):

    wb = load_workbook(filename=path)

    ws = wb.active

    # Progress Bar Object
    progsc = progressbar

    for i in progsc.progressbar(range(0, len(df.index))):

        url = df['image'][i]

        j = i + 2

        cellName = 'S{fnum}'.format(fnum=j)

        try:

            im = Image.open(requests.get(url, stream=True).raw)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(j)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

            wb.save(path)

            del cellName
            del url
            del im_100
            del im

            gc.collect()

        except requests.exceptions.MissingSchema:

            pass

        except requests.exceptions.ConnectionError:

            time.sleep(10)

            pass


# Image Downloader
def img_download_processing(df, path):

    ws = 0

    print('Initiating Multi-threading processing...')
    print('Image Download Only!')

    process_list = img_pre_multiprocess(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        p.map(img_multiprocess, process_list)

    # print(images[0][0][2]) # This the Path

    #img_post_multiprocess(images)

    print('MULTI-THREADING PROCESS ENDED')

    return True


# MAIN PROCESSES
def processing_xlsx(df, path):

    ws = 0

    print('Initiating Multi-threading processing...')

    process_list = img_pre_multiprocess(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        images=(p.map(img_multiprocess, process_list))

    # print(images[0][0][2]) # This the Path

    img_post_multiprocess(images)

    print('MULTI-THREADING PROCESS ENDED')

    return True


def processing_xlsx_local(df, path):

    print('Initiating Multi-threading processing...')
    print('Running Locally!')

    ws = 0

    process_list = img_pre_multiprocess(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        images=(p.map(img_multiprocess_local, process_list))

    # print(images[0][0][2]) # This the Path

    img_post_multiprocess(images)

    print('MULTI-THREADING PROCESS ENDED')

    return True


def processing_xlsx_local_inj(df, path, size):

    print('Initiating Multi-threading processing...')
    print('Running Locally!')

    ws = 0

    process_list = img_pre_multiprocess(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        images=(p.map(img_multiprocess_local, process_list))

    # print(images[0][0][2]) # This the Path

    img_post_multiprocess_inj(images, size)

    print('MULTI-THREADING PROCESS ENDED')

    return True


def processing_xlsx_inj(df, path, size):

    print('Initiating Multi-threading processing...')
    print('Downloading images!')

    ws = 0

    process_list = img_pre_multiprocess(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        images=(p.map(img_multiprocess, process_list))

    # print(images[0][0][2]) # This the Path

    img_post_multiprocess_inj(images, size)

    print('MULTI-THREADING PROCESS ENDED')

    return True


# PRE PROCESS
def img_pre_multiprocess(df, ws, path):

    time.sleep(4)

    # Progress Bar Object
    #progsc = progressbar

    cellprocess = []

    for i in (range(0, len(df.index))):

        url = df['image'][i]

        j = i + 2

        cellname = 'S{fnum}'.format(fnum=j)

        cellprocess.append([cellname, url, ws, path])

    return cellprocess


# POST PROCESSING
def img_post_multiprocess(images):

    #time.sleep(4)

    path = images[0][0][2]

    print(path)

    wb = load_workbook(filename=path)

    # Progress Bar Object
    progsq = progressbar

    for i in progsq.progressbar(range(0, len(images))):

        loc = images[i][0][0]

        try:
            cellname = images[i][0][1]
        except IndexError:
            continue

        try:

            imgd = Image.open(loc)

            xImg = pyImage(imgd)

            ws = wb.active

            ws[cellname] = ''

            ws.add_image(xImg, cellname)

        # This Exception is only raised when running local=True
        except FileNotFoundError:

            continue

        try:
            wb.save(path)
        except PermissionError:
            print('It seems like the xlsx is being used by another program, please close it first and try again.')

    return True


def img_post_multiprocess_inj(images, size=50):

    cyl = injector.cylinder(images, size)

    cyl.injection()

    return True


# COMPOSITEUR DEV


#Composer gets processes


#Download Images and Set Processes

def img_multiprocess(processes):

    cellName = processes[0]
    url = processes[1]
    path = processes[3]

    post_process_list = []

    try:

        im = Image.open(requests.get(url, stream=True).raw)

        im_100 = im.resize((78, 100))

        loc = "img/Sneaker{}.png".format(cellName)

        im_100.save(loc, format="png")

        #img = Image.open(loc)

        #xImg = pyImage(img)

        #ws[cellName] = ''

        #ws.add_image(xImg, cellName)

        post_process_list.append([loc, cellName, path])

        return post_process_list

    except requests.exceptions.MissingSchema:

        return 'The cell {fimg} is MissingSchema :('.format(fimg=cellName)

    except requests.exceptions.ConnectionError:

        return 'The cell {fimg} gives ConnectionError :('.format(fimg=cellName)


#Only Set Processes
def img_multiprocess_local(processes):

    cellName = processes[0]
    url = processes[1]
    path = processes[3]

    post_process_list = []

    try:

        #im = Image.open(requests.get(url, stream=True).raw)

        #im_100 = im.resize((78, 100))

        loc = "img/Sneaker{}.png".format(cellName)

        #im_100.save(loc, format="png")

        #img = Image.open(loc)

        #xImg = pyImage(img)

        #ws[cellName] = ''

        #ws.add_image(xImg, cellName)

        post_process_list.append([loc, cellName, path])

        return post_process_list

    except requests.exceptions.MissingSchema:

        return 'The cell {fimg} is MissingSchema :('.format(fimg=cellName)

    except requests.exceptions.ConnectionError:

        return 'The cell {fimg} gives ConnectionError :('.format(fimg=cellName)


# Composer sends to processor

#PROCESSORS
def img_processor(process_list, local=False):

    ws = 0

    print('Initiating Multi-threading processing...')
    print('Listing processes only')

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        if local == True:
            images = p.map(img_multiprocess_local, process_list)
        else:
            images = p.map(img_multiprocess, process_list)


    # print(images[0][0][2]) # This the Path

    img_post_multiprocess(images)

    print('MULTI-THREADING PROCESS ENDED')

    return images


def img_processor_inj(process_list, size, local):

    ws = 0

    print('Initiating Multi-threading processing...')


    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')

        print('Please wait...')
        if local == True:
            print('RUNNING LOCAL!')
            images = p.map(img_multiprocess_local, process_list)
        else:
            print('RUNNING ONLINE!')
            images = p.map(img_multiprocess, process_list)


    # print(images[0][0][2]) # This the Path

    img_post_multiprocess_inj(images, size)

    print('MULTI-THREADING PROCESS ENDED')

    return images

# ------------------------ DEV ----------------

def img_pre_multiprocess_training(df, ws, path):

    time.sleep(4)

    # Progress Bar Object
    #progsc = progressbar

    cellprocess = []

    for i in (range(0, len(df.index))):

        url = df['image'][df.index[i]]

        name = df['name'][df.index[i]]

        npath = path + '/' + name

        dataset.create_training_subfolder(name, path)

        j = i + 2

        cellname = 'S{fnum}'.format(fnum=j)

        cellprocess.append([cellname, url, ws, npath])


    return cellprocess

def img_multiprocess_training(processes):

    cellName = processes[0]
    url = processes[1]
    path = processes[3]

    post_process_list = []

    try:

        im = Image.open(requests.get(url, stream=True).raw)

        im_100 = im.resize((78, 100))

        loc = "img/Sneaker{}.png".format(cellName)

        nloc = path + '/1.png'


        im_100.save(nloc, format="png")

        #img = Image.open(loc)

        #xImg = pyImage(img)

        #ws[cellName] = ''

        #ws.add_image(xImg, cellName)

        post_process_list.append([loc, cellName, path])

        return post_process_list

    except requests.exceptions.MissingSchema:

        return 'The cell {fimg} is MissingSchema :('.format(fimg=cellName)

    except requests.exceptions.ConnectionError:

        return 'The cell {fimg} gives ConnectionError :('.format(fimg=cellName)

def img_downloader_training(df, path):

    ws = 0

    print('Initiating Multi-threading processing...')
    print('Image Download for Training Only!')

    process_list = img_pre_multiprocess_training(df, ws, path)

    with Pool(5) as p:

        #print(p.map(img_multiprocess, process_list))
        print('MULTI-THREADING PROCESS STARTED')
        print('Please wait...')
        p.map(img_multiprocess_training, process_list)

    # print(images[0][0][2]) # This the Path

    #img_post_multiprocess(images)

    print('MULTI-THREADING PROCESS ENDED')

    return True