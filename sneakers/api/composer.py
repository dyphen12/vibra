"""

Prisma Inc. 2021

composer.py

Status: Checked

Note: For unit testing purposes.

Made by Alexis W.

"""
import pandas as pd
import json
from openpyxl import load_workbook
from sneakers.api import utils
from sneakers.api import processing
from sneakers.api import core
from sneakers.api import uploaders


def chunks(data, n):

    m = int(len(data)/n)

    return [data[x:x+m] for x in range(0, len(data), m)]


# Composer is intended to work only on an Excel Worksheet object environment.
class Composer:
    def __init__(self, title):
        self.exist = False
        self.samplesize = 5
        self.folder = 'workout'
        self.title = title
        self.doc_file = '{ftit}.xlsx'.format(ftit=title)
        self.full_path = 'workout/{ftit}.xlsx'.format(ftit=title)
        self.doc_id = 'NaN'
        self.online = False
        self.json_name = 'workout/{}.json'.format(title)


        try:
            wb = load_workbook(filename=self.full_path)
            self.exist = True

        except FileNotFoundError as e:
            print('Composer workbook not found.')
            print('Creating workbook...')
            self.create_workbook()
            print('Workbook Created!')
            print('Check {}'.format(self.full_path))
            self.exist = True


        if self.exist == True:
            try:
                with open(self.json_name) as jsonFile:
                    jsonObject = json.load(jsonFile)
                    jsonFile.close()

                if jsonObject['composer']['doc_id'] == 'NaN':
                    self.online = False

                else:
                    self.doc_id = jsonObject['composer']['doc_id']
                    self.online = True

                self.samplesize = jsonObject['composer']['size']


            except FileNotFoundError:

                print('JSON does not exist')

    def create_workbook(self, sz=5):

        try:
            wb = load_workbook(filename=self.full_path)
            print('Workbook already exists!')
            self.exist = True

        except FileNotFoundError as e:
            # Creates a new empty workbook with no sneaker images, just data
            dataset = utils.load_shoes_dataset()
            sample = dataset.iloc[:sz]

            sample['pic'] = ''
            writer = pd.ExcelWriter(self.full_path, engine='xlsxwriter', options={'strings_to_urls': False})
            sample.to_excel(writer)
            writer.close()

            aid = 'NaN'

            json_comp = {
                "composer": {
                    "doc_id": 'NaN',
                    "size": sz
                }
            }

            with open('workout/{}.json'.format(self.title), 'w', encoding='utf-8') as f:

                json.dump(json_comp, f, ensure_ascii=False, indent=4)

            print('Workbook saved to ', self.full_path)


    def load_wb(self):
        wb = load_workbook(filename=self.full_path)
        return wb

    def write_wb(self, addr, local=False):

        cf = core.load_config()

        if cf.localimg == 'true':
            local=True

        if addr[1] > self.samplesize:
            print('Index Error, the max number of rows to image injection is {}.'.format(self.samplesize))
            print('Maybe you should expand the worksheet first.')
            return False


        msg = 'Writing from {ffrom} to {fto}'.format(ffrom=addr[0], fto=addr[1])
        print(msg)
        df = utils.load_shoes_dataset()
        sample = df.iloc[:self.samplesize]
        wb = self.load_wb()
        ws = wb.active
        processes = processing.img_pre_multiprocess(sample, ws, self.full_path)

        selection = []
        writings = []

        for i in range(addr[0], addr[1]+1):
            cellname = 'S{fnum}'.format(fnum=i)
            selection.append(cellname)

        for cell in selection:
            for i in range(0, len(processes)):
                if cell == processes[i][0]:
                    writings.append(processes[i])

        if local == True:
            processing.img_processor(writings, local=True)
        else:
            processing.img_processor(process_list=writings)

        print('XLSX processed successfully')

        return True

    def write_wb_xl(self, addr, iny_size=10):

        cf = core.load_config()

        if cf.localimg == 'true':
            local = True
        else:
            local = False

        if addr[1] > self.samplesize:
            print('Index Error, the max number of rows to image injection is {}.'.format(self.samplesize))
            print('Maybe you should expand the worksheet first.')

            return False
        msg = 'Writing from {ffrom} to {fto}'.format(ffrom=addr[0], fto=addr[1])
        print(msg)
        df = utils.load_shoes_dataset()
        sample = df.iloc[:self.samplesize]
        wb = self.load_wb()
        ws = wb.active
        processes = processing.img_pre_multiprocess(sample, ws, self.full_path)

        selection = []
        writings = []

        for i in range(addr[0], addr[1]+1):
            cellname = 'S{fnum}'.format(fnum=i)
            selection.append(cellname)

        for cell in selection:
            for i in range(0, len(processes)):
                if cell == processes[i][0]:
                    writings.append(processes[i])

        if local == True:
            processing.img_processor_inj(writings, iny_size, local=True)
        else:
            processing.img_processor_inj(process_list=writings, size=iny_size, local=False)

        print('XLSX processed successfully')

        return True

    # CONNECTS TO CORE
    def update_prices(self):

        wb = self.load_wb()
        ws = wb.active

        # iterate through excel and display data
        # iterate through excel and display data
        for i in range(0, self.samplesize):

            price_cell = 'C{fnum}'.format(fnum=i+2)
            id_cell = 'E{fnum}'.format(fnum=i + 2)

            id = ws[id_cell]
            sneaker = core.get_shoe_by_id(id)
            try:
                new_price = sneaker[0]['estimatedMarketValue']
            except IndexError:
                continue
            ws[price_cell] = new_price

        wb.save(self.full_path)
        try:
            wb.save(self.full_path)
        except PermissionError:
            print('It seems like the xlsx is being used by another program, please close it first and try again.')

        print('Prices Updated Successfully')
        return True

    def upload_file(self):

        # Just for the fist time


        if self.online == False:


            aid = uploaders.upload_folder(self.full_path, self.doc_file)
            self.doc_id = aid

            json_comp = {
                            "composer": {
                                "doc_id": aid,
                                "size": self.samplesize
                            }
                        }

            with open(self.json_name, 'w', encoding='utf-8') as f:

                json.dump(json_comp, f, ensure_ascii=False, indent=4)



            self.online = True


        else:

            print('This file is already online')

        return True

    def sync_file(self):

        if self.online == True:

            uploaders.sync_by_id(self.doc_id, self.full_path, self.doc_file)

        else:

            print('Sync failed, this workbook is not in the cloud.')
            print('You can upload it with composer.upload_file() .')

        return True

    def expand_worksheet(self, new_size):

        if new_size == self.samplesize:
            print('The new size is the same as the current size.')
            return False

        print('WARNING: THIS IS GOING TO CHANGE THE SAMPLESIZE OF THE WORKSHEET.\n')

        wb = self.load_wb()

        ws = wb.active

        dataset = utils.load_shoes_dataset()
        sample = dataset.iloc[:new_size]

        diff = new_size - self.samplesize

        for i in range(self.samplesize, new_size):

            brand_cell = 'B{fnum}'.format(fnum=i+2)
            price_cell = 'C{fnum}'.format(fnum=i+2)
            gender_cell = 'D{fnum}'.format(fnum=i+2)
            id_cell = 'E{fnum}'.format(fnum=i+2)
            image_cell = 'F{fnum}'.format(fnum=i+2)
            thumbnail_cell = 'G{fnum}'.format(fnum=i+2)
            link1_cell = 'H{fnum}'.format(fnum=i+2)
            link2_cell = 'I{fnum}'.format(fnum=i+2)
            link3_cell = 'J{fnum}'.format(fnum=i+2)
            name_cell = 'K{fnum}'.format(fnum=i+2)
            colorway_cell = 'L{fnum}'.format(fnum=i+2)
            releasedate_cell = 'M{fnum}'.format(fnum=i+2)
            releaseyear_cell = 'N{fnum}'.format(fnum=i+2)
            retailprices_cell = 'O{fnum}'.format(fnum=i+2)
            silhouette_cell = 'P{fnum}'.format(fnum=i+2)
            sku_cell = 'Q{fnum}'.format(fnum=i+2)
            story_cell = 'R{fnum}'.format(fnum=i+2)
            pic_cell = 'S{fnum}'.format(fnum=i+2)


            ws[brand_cell] = sample.iloc[i]['brand']
            ws[price_cell] = sample.iloc[i]['estimatedMarketValue']
            ws[gender_cell] = sample.iloc[i]['gender']
            ws[id_cell] = sample.iloc[i]['id']
            ws[image_cell] = sample.iloc[i]['image']
            ws[thumbnail_cell] = sample.iloc[i]['thumbnail']
            ws[link1_cell] = sample.iloc[i]['link1']
            ws[link2_cell] = sample.iloc[i]['link2']
            ws[link3_cell] = sample.iloc[i]['link3']
            ws[name_cell] = sample.iloc[i]['name']
            ws[colorway_cell] = sample.iloc[i]['colorway']
            ws[releasedate_cell] = sample.iloc[i]['releaseDate']
            ws[releaseyear_cell] = sample.iloc[i]['releaseYear']
            ws[retailprices_cell] = sample.iloc[i]['retailPrice']
            ws[silhouette_cell] = sample.iloc[i]['silhouette']
            ws[sku_cell] = sample.iloc[i]['sku']
            ws[story_cell] = sample.iloc[i]['story']
            ws[pic_cell] = ' '



        try:
            wb.save(self.full_path)
            print('{} rows added.'.format(diff))
        except PermissionError:
            print('It seems like the xlsx is being used by another program, please close it first and try again.')

        self.samplesize = new_size
        self.change_size_json(new_size)
        print('Worksheet Expanded Successfully')
        return

    def change_size_json(self, nsz):

        with open(self.json_name) as jsonFile:
            jsonObject = json.load(jsonFile)
            jsonFile.close()

        json_comp = {
            "composer": {
                "doc_id": jsonObject['composer']['doc_id'],
                "size": nsz
            }
        }

        with open(self.json_name, 'w', encoding='utf-8') as f:
            json.dump(json_comp, f, ensure_ascii=False, indent=4)



            # UPLOADERS FOR WEB

    def drive_flow_gui(self):

        url, ga = uploaders.get_drive_code()

        return url, ga
            # New Drive Algo

    def sync_worksheet(self, ga, code):

        if self.online is not True:

            aid = uploaders.upload_flow(ga, code, self.title, self.full_path)
            self.doc_id = aid

            json_comp = {
                "composer": {
                    "doc_id": aid,
                    "size": self.samplesize
                }
            }

            with open(self.json_name, 'w', encoding='utf-8') as f:

                json.dump(json_comp, f, ensure_ascii=False, indent=4)

            self.online = True

            print('Succesfully Uploaded.')
            return True
        else:
            uploaders.sync_flow(code, ga, self.doc_id, self.full_path, self.doc_file)
            print('Successfully Synced.')
            return True
























